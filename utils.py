"""
Вспомогательные функции для SERP Comparator
"""
import os
import shutil
import logging
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from constants import (
    SENTIMENT_COLORS_RU, 
    HEADER_FILL_COLOR, 
    HEADER_FONT_COLOR,
    DEFAULT_FONT_NAME
)

logger = logging.getLogger(__name__)


def cleanup_old_files(upload_folder, days=7):
    """Удаление файлов старше N дней"""
    try:
        cutoff_time = datetime.now() - timedelta(days=days)
        for root, dirs, files in os.walk(upload_folder):
            for dir_name in dirs:
                dir_path = os.path.join(root, dir_name)
                if os.path.getmtime(dir_path) < cutoff_time.timestamp():
                    shutil.rmtree(dir_path)
                    logger.info(f"Удалена старая папка: {dir_path}")
    except Exception as e:
        logger.error(f"Ошибка очистки файлов: {e}")


def create_excel_workbook(sheets_data):
    """
    Универсальная функция создания Excel с несколькими листами
    sheets_data: список кортежей (df, sheet_name) или (df, sheet_name, format_type)
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for i, sheet in enumerate(sheets_data):
            df = sheet[0]
            sheet_name = sheet[1]
            format_type = sheet[2] if len(sheet) > 2 else 'default'
            
            if df is not None and not df.empty:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                
                if format_type == 'url_stats':
                    format_url_stats_sheet(ws)
                elif format_type == 'comparison':
                    format_comparison_sheet(ws)
                elif format_type == 'summary':
                    format_summary_sheet(ws)
    
    output.seek(0)
    return output


def format_url_stats_sheet(ws):
    """Форматирование листа статистики URL"""
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    header_font = Font(name=DEFAULT_FONT_NAME, bold=True, color=HEADER_FONT_COLOR)
    
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 20
    for col_letter in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 15
    
    # Применяем заливку для URL в зависимости от тональности
    apply_sentiment_coloring(ws, sentiment_column=2, url_column=1)


def format_comparison_sheet(ws):
    """Форматирование листа сравнения"""
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    header_font = Font(name=DEFAULT_FONT_NAME, bold=True, color=HEADER_FONT_COLOR)
    
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.column_dimensions['B'].width = 60
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 15


def validate_file_path(requested_path, base_folder):
    """
    Безопасная валидация пути файла (защита от path traversal)
    Возвращает реальный путь если он находится в base_folder, иначе None
    """
    if not requested_path:
        return None
    
    try:
        # Если путь относительный, присоединяем к base_folder
        if not os.path.isabs(requested_path):
            # Убираем дублирующийся префикс 'user_data/' если он есть
            if requested_path.startswith('user_data/'):
                requested_path = requested_path[10:]
            requested_path = os.path.join(base_folder, requested_path)
        
        real_path = os.path.realpath(requested_path)
        base_real_path = os.path.realpath(base_folder)
        
        # Проверить, что путь находится внутри base_folder
        if real_path.startswith(base_real_path):
            return real_path
        return None
    except Exception:
        return None


def allowed_file(filename, allowed_extensions):
    """Проверка расширения файла"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


def get_file_size_mb(filepath):
    """Получить размер файла в МБ"""
    if os.path.exists(filepath):
        return os.path.getsize(filepath) / (1024 * 1024)
    return 0


def apply_sentiment_coloring(ws, sentiment_column=2, url_column=1):
    """
    Применяет заливку для URL в зависимости от тональности
    sentiment_column: колонка с тональностью
    url_column: колонка с URL для заливки
    """
    for row in range(2, ws.max_row + 1):
        sentiment = ws.cell(row=row, column=sentiment_column).value
        if sentiment in SENTIMENT_COLORS_RU:
            url_cell = ws.cell(row=row, column=url_column)
            fill = PatternFill(start_color=SENTIMENT_COLORS_RU[sentiment],
                             end_color=SENTIMENT_COLORS_RU[sentiment],
                             fill_type="solid")
            url_cell.fill = fill


def format_excel_headers(ws, column_count, font_name=None):
    """
    Форматирует заголовки Excel таблицы
    column_count: количество колонок для форматирования
    font_name: имя шрифта (по умолчанию из констант)
    """
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    header_font = Font(name=font_name or DEFAULT_FONT_NAME, bold=True, color=HEADER_FONT_COLOR)
    
    for col in range(1, column_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def set_column_widths(ws, widths_dict):
    """
    Устанавливает ширину колонок
    widths_dict: словарь {'A': 20, 'B': 30, ...}
    """
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def format_summary_sheet(ws):
    """Форматирование Summary листа"""
    format_excel_headers(ws, 3)
    set_column_widths(ws, {'A': 35, 'C': 100})
    
    # Автоширина для столбца B на основе содержимого
    max_length = 0
    for cell in ws['B']:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    # Устанавливаем ширину с небольшим отступом (минимум 12, максимум 40)
    adjusted_width = min(max(max_length + 2, 12), 40)
    ws.column_dimensions['B'].width = adjusted_width
    
    # Базовое форматирование для всех строк
    for row in range(2, ws.max_row + 1):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            
            # Добавляем границы
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = thin_border
            
            # Форматируем метрики жирным с эмодзи
            if col == 1 and cell.value:
                cell.font = Font(name=DEFAULT_FONT_NAME, bold=True, size=12)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                # Разный цвет фона для категорий
                if any(emoji in str(cell.value) for emoji in ['🔗', '📤', '📈', '📉', '🔄']):
                    if '🔗' in str(cell.value):
                        fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                    elif '📤' in str(cell.value):
                        fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
                    elif '📈' in str(cell.value):
                        fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                    elif '📉' in str(cell.value):
                        fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
                    elif '🔄' in str(cell.value):
                        fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    cell.fill = fill
            
            # Центрируем значения в столбце B без цветной заливки
            elif col == 2 and isinstance(cell.value, (int, float)):
                cell.font = Font(name=DEFAULT_FONT_NAME, bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Форматируем детали и URL
            elif col == 3:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                cell.font = Font(name=DEFAULT_FONT_NAME, size=10)
                
                # Светло-серый фон для строк с URL
                if cell.value and isinstance(cell.value, str):
                    if 'http' in cell.value or '•' in cell.value:
                        fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                        cell.fill = fill


def is_real_query(row):
    """Проверяет, является ли строка реальным поисковым запросом (не статистикой)"""
    query = row.get('Запрос')
    if pd.isna(query) or query == '':
        return False
    query_str = str(query).upper()
    if 'СТАТИСТИКА' in query_str:
        return False
    url = row.get('URL')
    if pd.isna(url) or url == '':
        return False
    return True


def is_valid_for_second_file(row, label2):
    """Проверяет, есть ли данные для второго файла (позиция и тональность)"""
    pos2_col = f'Позиция_{label2}'
    if pos2_col not in row.index or pd.isna(row[pos2_col]):
        return False
    sentiment_col = f'Тональность_{label2}'
    if sentiment_col not in row.index or pd.isna(row[sentiment_col]):
        return False
    sentiment = row[sentiment_col]
    if sentiment == "Неопределенная":
        return False
    return True
